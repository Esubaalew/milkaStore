from django.contrib import admin
from unfold.admin import ModelAdmin
from django.utils.html import format_html
from .models import Order, Product, Category, Subcategory, Brand, ProductModel, Stock, Purchase, Telegram
from django.core.exceptions import ValidationError
from django.http import HttpResponse
import csv
from openpyxl import Workbook


class OrderAdmin(ModelAdmin):
    model = Order
    list_display = (
        'product_name',
        'full_name',
        'address',
        'phone_number',
        'comment',
        'quantity',
        'order_date',
        'is_paid',
    )
    list_filter = (
        'product',
        'full_name',
        'address',
        'phone_number',
        'quantity',
        'order_date',
        'is_paid',
    )
    search_fields = (
        'product__name',
        'full_name',
        'address',
        'phone_number',
        'comment'
    )
    ordering = ('-order_date',)

    def product_name(self, obj):
        return obj.product.name
    product_name.short_description = 'Product'

    def export_as_csv(self, request, queryset):
        # Create the response object for CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="orders.csv"'

        writer = csv.writer(response)
        writer.writerow(['Product', 'Full Name', 'Address', 'Phone Number', 'Comment', 'Quantity', 'Order Date', 'Is Paid'])  # CSV Header

        for order in queryset:
            writer.writerow([
                order.product.name,
                order.full_name,
                order.address,
                order.phone_number,
                order.comment,
                order.quantity,
                order.order_date,
                order.is_paid,
            ])  # CSV Data

        return response

    def export_as_excel(self, request, queryset):
        # Create the response object for Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Orders'

        # Write headers
        headers = ['Product', 'Full Name', 'Address', 'Phone Number', 'Comment', 'Quantity', 'Order Date', 'Is Paid']
        worksheet.append(headers)

        # Write data
        for order in queryset:
            order_date_naive = order.order_date.replace(tzinfo=None)  # Make datetime naive
            worksheet.append([
                order.product.name,
                order.full_name,
                order.address,
                order.phone_number,
                order.comment,
                order.quantity,
                order_date_naive,  # Use the naive datetime
                order.is_paid,
            ])

        workbook.save(response)
        return response

    export_as_csv.short_description = "Export Selected Orders as CSV"
    export_as_excel.short_description = "Export Selected Orders as Excel"

    # Adding the actions to the admin
    actions = [export_as_csv, export_as_excel]

    def save_model(self, request, obj, form, change):
        try:
            super().save_model(request, obj, form, change)
        except ValidationError as e:
            form.add_error(None, e)
            raise

class ProductAdmin(ModelAdmin):
    model = Product
    list_display = (
        'code',
        'name',
        'short_description',
        'quantity',
        'image_preview',
        'price',
        'date_added'
    )
    list_filter = ('name', 'price', 'date_added')
    search_fields = ('code','name', 'price', 'date_added')
    ordering = ('-date_added',)

    def short_description(self, obj):
        return obj.description if len(obj.description) <= 50 else f"{obj.description[:50]}..."
    short_description.short_description = 'Description'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="width: 50px; height: auto;" />', obj.image.url)
        return "No image"
    image_preview.short_description = 'Image Preview'

    def export_as_csv(self, request, queryset):
        # Create the response object for CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'

        writer = csv.writer(response)
        writer.writerow(['Code', 'Name', 'Description', 'Quantity', 'Price', 'Date Added'])  # CSV Header

        for product in queryset:
            writer.writerow([product.code,product.name, product.description, product.quantity, product.price, product.date_added])  # CSV Data

        return response

    def export_as_excel(self, request, queryset):
        # Create the response object for Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Products'

        # Write headers
        headers = ['Name', 'Description', 'Quantity', 'Price', 'Date Added']
        worksheet.append(headers)

        # Write data
        for product in queryset:
            date_added_naive = product.date_added.replace(tzinfo=None)  # Make datetime naive
            worksheet.append([product.name, product.description, product.quantity, product.price, date_added_naive])  # Use the naive datetime

        workbook.save(response)
        return response

    export_as_csv.short_description = "Export Selected Products as CSV"
    export_as_excel.short_description = "Export Selected Products as Excel"

    # Adding the actions to the admin
    actions = [export_as_csv, export_as_excel]

class PurchaseAdmin(ModelAdmin):
    model = Purchase
    list_display = ('product_name', 'product_code', 'product_brand', 'quantity_purchased', 'purchase_date')
    list_filter = ('product__name', 'quantity_purchased', 'purchase_date', 'product__brand__name')  # Filter by product's brand
    search_fields = ('product__name', 'product__code', 'product__brand__name', 'quantity_purchased', 'purchase_date')

    # Custom method to display the product's name
    def product_name(self, obj):
        return obj.product.name
    product_name.short_description = 'Product Name'

    # Custom method to display the product's code
    def product_code(self, obj):
        return obj.product.code
    product_code.short_description = 'Product Code'

    # Custom method to display the product's brand
    def product_brand(self, obj):
        return obj.product.brand.name if obj.product.brand else None
    product_brand.short_description = 'Product Brand'

    # CSV export action
    def export_as_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="purchases.csv"'

        writer = csv.writer(response)
        writer.writerow(['Product Name', 'Product Code', 'Product Brand', 'Quantity Purchased', 'Purchase Date'])  # CSV Header

        for purchase in queryset:
            writer.writerow([
                purchase.product.name,
                purchase.product.code,
                purchase.product.brand.name if purchase.product.brand else '',
                purchase.quantity_purchased,
                purchase.purchase_date
            ])  # CSV Data

        return response

    # Excel export action
    def export_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="purchases.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Purchases'

        # Write headers
        headers = ['Product Name', 'Product Code', 'Product Brand', 'Quantity Purchased', 'Purchase Date']
        worksheet.append(headers)

        # Write data
        for purchase in queryset:
            worksheet.append([
                purchase.product.name,
                purchase.product.code,
                purchase.product.brand.name if purchase.product.brand else '',
                purchase.quantity_purchased,
                purchase.purchase_date
            ])  # Purchase data

        workbook.save(response)
        return response

    # Adding the actions to the admin
    actions = [export_as_csv, export_as_excel]

    export_as_csv.short_description = "Export Selected Purchases as CSV"
    export_as_excel.short_description = "Export Selected Purchases as Excel"

class CategoryAdmin(ModelAdmin):
    model = Category
    list_display = ('name', )
    list_filter = ('name', )
    search_fields = ('name', )

class SubcategoryAdmin(ModelAdmin):
    model = Subcategory
    list_display = ('name', 'category', )
    list_filter = ('name', 'category', )
    search_fields = ('name', 'category',)

class BrandAdmin(ModelAdmin):
    model = Brand
    list_display = ('name', 'subcategory', )
    list_filter = ('name', 'subcategory', )
    search_fields = ('name', 'subcategory',)

class ProductModelAdmin(ModelAdmin):
    model = ProductModel
    list_display = ('name', 'brand', 'subcategory', )
    list_filter = ('name', 'brand', 'subcategory', )
    search_fields = ('name', 'brand', 'subcategory',)

class StockAdmin(ModelAdmin):
    model = Stock
    list_display = ('product', 'quantity_in_stock', 'restock_date', 'added_by')  # Display the fields we have
    list_filter = ('product', 'quantity_in_stock', 'restock_date', 'added_by')  # Filterable fields
    search_fields = ('product__name', 'quantity_in_stock', 'restock_date')  # Searchable fields

    actions = ['restock_items']

    def restock_items(self, request, queryset):
        for stock in queryset:
            stock.quantity_in_stock += 10
            stock.save()

    restock_items.short_description = "Restock selected items"

    def save_model(self, request, obj, form, change):
        if not obj.added_by:
            obj.added_by = request.user

        obj.clean()
        super().save_model(request, obj, form, change)

class TelegramAdmin(ModelAdmin):
    model = Telegram
    list_display = ('product_name', 'product_code', 'quantity_in_stock', 'date_posted')
    search_fields = ('stock__product__name', 'stock__product__code', 'date_posted')

    def product_name(self, obj):
        return obj.stock.product.name
    product_name.short_description = 'Product Name'

    def product_code(self, obj):
        return obj.stock.product.code
    product_code.short_description = 'Product Code'

    def quantity_in_stock(self, obj):
        return obj.stock.quantity_in_stock
    quantity_in_stock.short_description = 'Quantity in Stock'

admin.site.register(Order, OrderAdmin)
admin.site.register(Product, ProductAdmin)
admin.site.register(Category, CategoryAdmin)
admin.site.register(Subcategory, SubcategoryAdmin)
admin.site.register(Brand, BrandAdmin)
admin.site.register(ProductModel, ProductModelAdmin)
admin.site.register(Stock, StockAdmin)
admin.site.register(Purchase, PurchaseAdmin)
admin.site.register(Telegram, TelegramAdmin)

admin.site.site_header = "Store Administration"
admin.site.site_title = "Shop Admin Portal"
admin.site.index_title = "Welcome to the Shop Admin Dashboard"
